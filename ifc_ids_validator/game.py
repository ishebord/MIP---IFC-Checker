# game.py
from __future__ import annotations

import math
import os
import random
import sys
import getpass
from dataclasses import dataclass
from pathlib import Path

import pygame
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


# ============================================================
# ОБЩИЕ НАСТРОЙКИ
# ============================================================

WIDTH, HEIGHT = 960, 560
FPS = 120

RESULT_DIR = Path(r"T:\BIM отдел\Шебордаев\programming\game_result")
RESULT_FILE = RESULT_DIR / "leaderboard.xlsx"

BG = (8, 12, 24)
WHITE = (245, 248, 255)
TEXT = (225, 235, 255)
MUTED = (125, 150, 190)
CYAN = (54, 220, 255)
BLUE = (60, 120, 255)
PURPLE = (165, 90, 255)
PINK = (255, 70, 165)
RED = (255, 70, 85)
ORANGE = (255, 165, 70)
GREEN = (70, 245, 160)
YELLOW = (255, 230, 90)


def clamp(v, a, b):
    return max(a, min(b, v))


def length(x, y):
    return math.hypot(x, y)


def normalize(x, y):
    l = math.hypot(x, y)
    if l <= 0.0001:
        return 0, 0
    return x / l, y / l


def lerp(a, b, t):
    return a + (b - a) * t


def draw_text(surface, font, text, x, y, color=TEXT, center=False):
    img = font.render(text, True, color)
    r = img.get_rect()
    if center:
        r.center = (x, y)
    else:
        r.topleft = (x, y)
    surface.blit(img, r)
    return r


def get_windows_user():
    return os.environ.get("USERNAME") or getpass.getuser() or "Unknown"


def load_leaderboard():
    try:
        if not RESULT_FILE.exists():
            return []

        wb = load_workbook(RESULT_FILE)
        ws = wb.active

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue

            user = str(row[0])
            score = int(row[1] or 0)
            rows.append((user, score))

        rows.sort(key=lambda x: x[1], reverse=True)
        return rows

    except Exception:
        return []


def save_score_to_excel(user, score):
    RESULT_DIR.mkdir(parents=True, exist_ok=True)

    if RESULT_FILE.exists():
        wb = load_workbook(RESULT_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Лидеры"
        ws["A1"] = "Пользователь"
        ws["B1"] = "Максимальные очки"

    existing_row = None
    existing_best = 0

    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if str(name).lower() == str(user).lower():
            existing_row = row
            existing_best = int(ws.cell(row=row, column=2).value or 0)
            break

    if existing_row:
        if score > existing_best:
            ws.cell(row=existing_row, column=2).value = score
    else:
        new_row = ws.max_row + 1
        ws.cell(row=new_row, column=1).value = user
        ws.cell(row=new_row, column=2).value = score

    data = []
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        value = ws.cell(row=row, column=2).value
        if name:
            data.append((str(name), int(value or 0)))

    data.sort(key=lambda x: x[1], reverse=True)

    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row)

    for i, (name, value) in enumerate(data, start=2):
        ws.cell(row=i, column=1).value = name
        ws.cell(row=i, column=2).value = value

    header_fill = PatternFill("solid", fgColor="0B4EA2")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style="thin", color="D6DFEB"),
        right=Side(style="thin", color="D6DFEB"),
        top=Side(style="thin", color="D6DFEB"),
        bottom=Side(style="thin", color="D6DFEB")
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if cell.column == 2 else "left")

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22
    ws.freeze_panes = "A2"

    wb.save(RESULT_FILE)


# ============================================================
# ЧАСТИЦЫ
# ============================================================

@dataclass
class Particle:
    x: float
    y: float
    vx: float
    vy: float
    radius: float
    color: tuple[int, int, int]
    life: float
    max_life: float
    glow: bool = True
    gravity: float = 0.0

    def update(self, dt):
        self.life -= dt
        self.vy += self.gravity * dt
        self.x += self.vx * dt
        self.y += self.vy * dt
        self.vx *= 0.985
        self.vy *= 0.985
        return self.life > 0

    def draw(self, surf):
        k = clamp(self.life / self.max_life, 0, 1)
        r = max(1, int(self.radius * k))
        color = tuple(int(c * k) for c in self.color)
        if self.glow:
            pygame.draw.circle(surf, tuple(int(c * 0.35) for c in color), (int(self.x), int(self.y)), r * 4)
        pygame.draw.circle(surf, color, (int(self.x), int(self.y)), r)


# ============================================================
# NEON CORE
# ============================================================

class Player:
    def __init__(self):
        self.x = WIDTH / 2
        self.y = HEIGHT / 2
        self.vx = 0
        self.vy = 0
        self.radius = 17
        self.speed = 520
        self.accel = 13
        self.hp = 100
        self.max_hp = 100
        self.energy = 100
        self.max_energy = 100
        self.dash_cd = 0
        self.invuln = 0
        self.angle = 0

    def update(self, dt, keys, mouse_buttons):
        ix = int(keys[pygame.K_d] or keys[pygame.K_RIGHT]) - int(keys[pygame.K_a] or keys[pygame.K_LEFT])
        iy = int(keys[pygame.K_s] or keys[pygame.K_DOWN]) - int(keys[pygame.K_w] or keys[pygame.K_UP])
        nx, ny = normalize(ix, iy)

        self.vx = lerp(self.vx, nx * self.speed, min(1, self.accel * dt))
        self.vy = lerp(self.vy, ny * self.speed, min(1, self.accel * dt))

        self.x += self.vx * dt
        self.y += self.vy * dt

        self.x = clamp(self.x, 30, WIDTH - 30)
        self.y = clamp(self.y, 30, HEIGHT - 30)

        self.energy = min(self.max_energy, self.energy + 22 * dt)
        self.dash_cd = max(0, self.dash_cd - dt)
        self.invuln = max(0, self.invuln - dt)

        if length(self.vx, self.vy) > 30:
            self.angle = math.atan2(self.vy, self.vx)

    def dash(self, particles):
        if self.dash_cd > 0 or self.energy < 32:
            return

        mx, my = pygame.mouse.get_pos()
        dx, dy = normalize(mx - self.x, my - self.y)

        if dx == 0 and dy == 0:
            dx, dy = math.cos(self.angle), math.sin(self.angle)

        self.x += dx * 115
        self.y += dy * 115
        self.x = clamp(self.x, 30, WIDTH - 30)
        self.y = clamp(self.y, 30, HEIGHT - 30)

        self.energy -= 32
        self.dash_cd = 0.42
        self.invuln = 0.28

        for _ in range(34):
            a = math.atan2(-dy, -dx) + random.uniform(-0.75, 0.75)
            s = random.uniform(90, 480)
            particles.append(Particle(
                self.x, self.y,
                math.cos(a) * s,
                math.sin(a) * s,
                random.uniform(2, 6),
                random.choice([CYAN, BLUE, PURPLE]),
                random.uniform(0.18, 0.45),
                0.45
            ))

    def damage(self, amount, particles):
        if self.invuln > 0:
            return False

        self.hp -= amount
        self.invuln = 0.65

        for _ in range(28):
            a = random.uniform(0, math.tau)
            s = random.uniform(80, 360)
            particles.append(Particle(
                self.x, self.y,
                math.cos(a) * s,
                math.sin(a) * s,
                random.uniform(2, 7),
                random.choice([RED, PINK, ORANGE]),
                random.uniform(0.25, 0.6),
                0.6
            ))
        return True

    def draw(self, surf, t):
        pulse = 1 + math.sin(t * 8) * 0.04
        r = int(self.radius * pulse)

        if self.invuln > 0:
            pygame.draw.circle(surf, (80, 190, 255), (int(self.x), int(self.y)), r + 12, 2)

        pygame.draw.circle(surf, (20, 80, 150), (int(self.x), int(self.y)), r + 12)
        pygame.draw.circle(surf, CYAN, (int(self.x), int(self.y)), r)
        pygame.draw.circle(surf, WHITE, (int(self.x - 5), int(self.y - 6)), 5)

        nose_x = self.x + math.cos(self.angle) * 25
        nose_y = self.y + math.sin(self.angle) * 25
        pygame.draw.line(surf, WHITE, (self.x, self.y), (nose_x, nose_y), 4)


class Enemy:
    def __init__(self, kind, x, y):
        self.kind = kind
        self.x = x
        self.y = y
        self.vx = 0
        self.vy = 0
        self.dead = False
        self.flash = 0

        if kind == "chaser":
            self.radius = 15
            self.speed = 125
            self.hp = 20
            self.color = RED
            self.value = 12
        elif kind == "speeder":
            self.radius = 11
            self.speed = 235
            self.hp = 10
            self.color = ORANGE
            self.value = 18
        elif kind == "tank":
            self.radius = 26
            self.speed = 72
            self.hp = 55
            self.color = PURPLE
            self.value = 35
        else:
            self.radius = 14
            self.speed = 110
            self.hp = 15
            self.color = PINK
            self.value = 15

    def update(self, dt, player, enemies):
        dx, dy = normalize(player.x - self.x, player.y - self.y)

        wobble = math.sin((self.x + self.y) * 0.02 + pygame.time.get_ticks() * 0.006) * 0.8
        tx = dx * math.cos(wobble) - dy * math.sin(wobble)
        ty = dx * math.sin(wobble) + dy * math.cos(wobble)

        self.vx = lerp(self.vx, tx * self.speed, min(1, 5 * dt))
        self.vy = lerp(self.vy, ty * self.speed, min(1, 5 * dt))

        self.x += self.vx * dt
        self.y += self.vy * dt

        for other in enemies:
            if other is self or other.dead:
                continue
            dxo = self.x - other.x
            dyo = self.y - other.y
            d = length(dxo, dyo)
            min_d = self.radius + other.radius + 4
            if 0 < d < min_d:
                nx, ny = dxo / d, dyo / d
                push = (min_d - d) * 0.5
                self.x += nx * push
                self.y += ny * push

        self.flash = max(0, self.flash - dt)

    def hit(self, damage):
        self.hp -= damage
        self.flash = 0.08
        if self.hp <= 0:
            self.dead = True

    def draw(self, surf, t):
        color = WHITE if self.flash > 0 else self.color
        x, y = int(self.x), int(self.y)

        pygame.draw.circle(surf, tuple(int(c * 0.25) for c in color), (x, y), self.radius * 3)
        pygame.draw.circle(surf, color, (x, y), self.radius)

        if self.kind == "tank":
            pygame.draw.circle(surf, (40, 20, 70), (x, y), self.radius - 8)
        elif self.kind == "speeder":
            pygame.draw.polygon(
                surf,
                WHITE,
                [
                    (x + math.cos(t * 8) * 4, y - 5),
                    (x + 7, y + 5),
                    (x - 7, y + 5),
                ]
            )
        else:
            pygame.draw.circle(surf, (30, 10, 20), (x - 5, y - 3), 3)
            pygame.draw.circle(surf, (30, 10, 20), (x + 5, y - 3), 3)


class Orb:
    def __init__(self, x, y):
        self.x = x
        self.y = y
        self.r = 8
        self.phase = random.random() * math.tau
        self.dead = False

    def update(self, dt, player):
        d = length(player.x - self.x, player.y - self.y)
        if d < 135:
            nx, ny = normalize(player.x - self.x, player.y - self.y)
            self.x += nx * 360 * dt
            self.y += ny * 360 * dt

        if d < player.radius + self.r + 8:
            self.dead = True
            return True
        return False

    def draw(self, surf, t):
        r = self.r + math.sin(t * 8 + self.phase) * 2
        pygame.draw.circle(surf, (30, 90, 70), (int(self.x), int(self.y)), int(r * 3))
        pygame.draw.circle(surf, GREEN, (int(self.x), int(self.y)), int(r))
        pygame.draw.circle(surf, WHITE, (int(self.x - 2), int(self.y - 2)), 2)


# ============================================================
# TETRIS
# ============================================================

TETRIS_COLS = 10
TETRIS_ROWS = 20
TETRIS_CELL = 24
TETRIS_X = 258
TETRIS_Y = 42

TETRIS_SHAPES = {
    "I": [[1, 1, 1, 1]],
    "O": [[1, 1], [1, 1]],
    "T": [[0, 1, 0], [1, 1, 1]],
    "S": [[0, 1, 1], [1, 1, 0]],
    "Z": [[1, 1, 0], [0, 1, 1]],
    "J": [[1, 0, 0], [1, 1, 1]],
    "L": [[0, 0, 1], [1, 1, 1]],
}

TETRIS_COLORS = {
    "I": (53, 220, 255),
    "O": (255, 230, 90),
    "T": (165, 90, 255),
    "S": (70, 245, 160),
    "Z": (255, 70, 85),
    "J": (60, 120, 255),
    "L": (255, 165, 70),
}


class TetrisPiece:
    def __init__(self, kind=None):
        self.kind = kind or random.choice(list(TETRIS_SHAPES.keys()))
        self.matrix = [row[:] for row in TETRIS_SHAPES[self.kind]]
        self.x = TETRIS_COLS // 2 - len(self.matrix[0]) // 2
        self.y = -1
        self.color = TETRIS_COLORS[self.kind]

    def rotated(self):
        rows = len(self.matrix)
        cols = len(self.matrix[0])
        return [[self.matrix[rows - 1 - r][c] for r in range(rows)] for c in range(cols)]


class TetrisGame:
    def __init__(self, outer: "Game"):
        self.outer = outer
        self.font_big = outer.font_big
        self.font_mid = outer.font_mid
        self.font = outer.font
        self.font_small = outer.font_small
        self.reset()

    def reset(self):
        self.board = [[None for _ in range(TETRIS_COLS)] for _ in range(TETRIS_ROWS)]
        self.bag = []
        self.current = self.next_piece()
        self.next_items = [self.next_piece() for _ in range(3)]
        self.hold = None
        self.can_hold = True

        self.score = 0
        self.lines = 0
        self.level = 1
        self.combo = 0
        self.drop_timer = 0
        self.lock_timer = 0
        self.soft_drop = False
        self.game_over = False
        self.pause = False

        # Мягкое управление по горизонтали (DAS/ARR как в хороших тетрисах):
        # первое нажатие двигает на 1 клетку, затем есть небольшая пауза,
        # и только потом начинается спокойный автоповтор.
        self.move_cooldown = 0
        self.move_direction = 0
        self.move_das = 0.0
        self.move_arr = 0.0
        self.move_initial_delay = 0.18   # задержка перед автоповтором
        self.move_repeat_delay = 0.085   # частота автоповтора

        # Отдельные координаты только для отрисовки — фигура визуально
        # плавно догоняет логическую позицию, а не дёргается мгновенно.
        self.visual_x = float(self.current.x)
        self.visual_y = float(self.current.y)

        self.rotate_cooldown = 0
        self.hard_drop_flash = 0
        self.line_flash = []
        self.particles: list[Particle] = []
        self.message = ""
        self.message_timer = 0
        self.t = 0

    def next_piece(self):
        if not self.bag:
            self.bag = list(TETRIS_SHAPES.keys())
            random.shuffle(self.bag)
        return TetrisPiece(self.bag.pop())

    def spawn_next(self):
        self.current = self.next_items.pop(0)
        self.next_items.append(self.next_piece())
        self.current.x = TETRIS_COLS // 2 - len(self.current.matrix[0]) // 2
        self.current.y = -1
        self.visual_x = float(self.current.x)
        self.visual_y = float(self.current.y)
        self.can_hold = True
        if not self.valid(self.current.x, self.current.y, self.current.matrix):
            self.game_over = True
            self.message = "СТЕК ПЕРЕПОЛНЕН"
            self.message_timer = 999

    def valid(self, x, y, matrix):
        for r, row in enumerate(matrix):
            for c, val in enumerate(row):
                if not val:
                    continue
                bx = x + c
                by = y + r
                if bx < 0 or bx >= TETRIS_COLS or by >= TETRIS_ROWS:
                    return False
                if by >= 0 and self.board[by][bx] is not None:
                    return False
        return True

    def move(self, dx, dy):
        if self.valid(self.current.x + dx, self.current.y + dy, self.current.matrix):
            self.current.x += dx
            self.current.y += dy
            return True
        return False

    def rotate(self, direction=1):
        mat = self.current.rotated()
        if direction < 0:
            # Три поворота по часовой = один против часовой.
            for _ in range(2):
                rows = len(mat)
                cols = len(mat[0])
                mat = [[mat[rows - 1 - r][c] for r in range(rows)] for c in range(cols)]

        for kick in [0, -1, 1, -2, 2]:
            if self.valid(self.current.x + kick, self.current.y, mat):
                self.current.x += kick
                self.current.matrix = mat
                self.spawn_piece_spark()
                return True
        return False

    def hold_piece(self):
        if not self.can_hold:
            return
        old = self.current
        if self.hold is None:
            self.hold = TetrisPiece(old.kind)
            self.spawn_next()
        else:
            self.current = TetrisPiece(self.hold.kind)
            self.hold = TetrisPiece(old.kind)
            self.visual_x = float(self.current.x)
            self.visual_y = float(self.current.y)
        self.can_hold = False
        self.spawn_piece_spark()

    def hard_drop(self):
        dist = 0
        while self.move(0, 1):
            dist += 1
        self.score += dist * 2
        self.visual_x = float(self.current.x)
        self.visual_y = float(self.current.y)
        self.hard_drop_flash = 0.16
        self.lock_piece()

    def ghost_y(self):
        gy = self.current.y
        while self.valid(self.current.x, gy + 1, self.current.matrix):
            gy += 1
        return gy

    def lock_piece(self):
        for r, row in enumerate(self.current.matrix):
            for c, val in enumerate(row):
                if not val:
                    continue
                bx = self.current.x + c
                by = self.current.y + r
                if by < 0:
                    self.game_over = True
                    return
                self.board[by][bx] = self.current.kind
                self.spawn_block_particles(bx, by, self.current.color, count=4)

        self.clear_lines()
        self.spawn_next()

    def clear_lines(self):
        full = [i for i, row in enumerate(self.board) if all(cell is not None for cell in row)]
        if not full:
            self.combo = 0
            return

        self.line_flash = [(row, 0.28) for row in full]

        for row_i in full:
            for col in range(TETRIS_COLS):
                kind = self.board[row_i][col]
                color = TETRIS_COLORS.get(kind, WHITE)
                self.spawn_block_particles(col, row_i, color, count=9)

        new_board = [row for i, row in enumerate(self.board) if i not in full]
        while len(new_board) < TETRIS_ROWS:
            new_board.insert(0, [None for _ in range(TETRIS_COLS)])
        self.board = new_board

        n = len(full)
        self.lines += n
        self.level = 1 + self.lines // 10
        self.combo += 1

        points = {1: 100, 2: 300, 3: 500, 4: 900}.get(n, 1200)
        self.score += points * self.level + self.combo * 45

        if n == 4:
            self.message = "TETRIS!"
            self.message_timer = 1.2
        elif self.combo > 2:
            self.message = f"COMBO x{self.combo}"
            self.message_timer = 0.9
        else:
            self.message = f"+{points * self.level}"
            self.message_timer = 0.7

    def spawn_piece_spark(self):
        color = self.current.color
        cx = TETRIS_X + (self.current.x + 1.5) * TETRIS_CELL
        cy = TETRIS_Y + (self.current.y + 1.5) * TETRIS_CELL
        for _ in range(8):
            a = random.uniform(0, math.tau)
            s = random.uniform(35, 130)
            self.particles.append(Particle(
                cx, cy, math.cos(a) * s, math.sin(a) * s,
                random.uniform(1.5, 3.5), color, 0.25, 0.25
            ))

    def spawn_block_particles(self, col, row, color, count=5):
        x = TETRIS_X + col * TETRIS_CELL + TETRIS_CELL / 2
        y = TETRIS_Y + row * TETRIS_CELL + TETRIS_CELL / 2
        for _ in range(count):
            a = random.uniform(0, math.tau)
            s = random.uniform(35, 190)
            self.particles.append(Particle(
                x, y, math.cos(a) * s, math.sin(a) * s,
                random.uniform(1.5, 4.2), color,
                random.uniform(0.25, 0.65), 0.65, gravity=160
            ))

    def fall_delay(self):
        return max(0.065, 0.72 * (0.86 ** (self.level - 1)))

    def handle_event(self, event):
        if event.type == pygame.KEYDOWN:
            if event.key == pygame.K_ESCAPE:
                self.outer.state = "select"
                return

            if self.game_over:
                if event.key in (pygame.K_RETURN, pygame.K_SPACE):
                    self.reset()
                return

            if event.key == pygame.K_p:
                self.pause = not self.pause
            if self.pause:
                return

            if event.key in (pygame.K_LEFT, pygame.K_a):
                self.move_direction = -1
                self.move_das = self.move_initial_delay
                self.move_arr = 0.0
                self.move(-1, 0)
            elif event.key in (pygame.K_RIGHT, pygame.K_d):
                self.move_direction = 1
                self.move_das = self.move_initial_delay
                self.move_arr = 0.0
                self.move(1, 0)
            elif event.key in (pygame.K_DOWN, pygame.K_s):
                self.soft_drop = True
            elif event.key in (pygame.K_UP, pygame.K_w, pygame.K_x):
                self.rotate(1)
            elif event.key in (pygame.K_z,):
                self.rotate(-1)
            elif event.key == pygame.K_SPACE:
                self.hard_drop()
            elif event.key in (pygame.K_c, pygame.K_LSHIFT, pygame.K_RSHIFT):
                self.hold_piece()

        if event.type == pygame.KEYUP:
            if event.key in (pygame.K_DOWN, pygame.K_s):
                self.soft_drop = False
            elif event.key in (pygame.K_LEFT, pygame.K_a) and self.move_direction == -1:
                self.move_direction = 0
                self.move_das = 0.0
                self.move_arr = 0.0
            elif event.key in (pygame.K_RIGHT, pygame.K_d) and self.move_direction == 1:
                self.move_direction = 0
                self.move_das = 0.0
                self.move_arr = 0.0

    def update(self, dt):
        self.t += dt
        self.message_timer = max(0, self.message_timer - dt)
        self.hard_drop_flash = max(0, self.hard_drop_flash - dt)
        self.move_cooldown = max(0, self.move_cooldown - dt)
        self.rotate_cooldown = max(0, self.rotate_cooldown - dt)

        for p in list(self.particles):
            if not p.update(dt):
                self.particles.remove(p)

        self.line_flash = [(row, t - dt) for row, t in self.line_flash if t - dt > 0]

        if self.game_over or self.pause:
            return

        keys = pygame.key.get_pressed()

        # Мягкий автоповтор по горизонтали.
        # Нажатие = 1 шаг. Удержание = пауза DAS, затем редкие повторные шаги ARR.
        if self.move_direction:
            still_holding_left = keys[pygame.K_LEFT] or keys[pygame.K_a]
            still_holding_right = keys[pygame.K_RIGHT] or keys[pygame.K_d]
            if (self.move_direction == -1 and not still_holding_left) or (self.move_direction == 1 and not still_holding_right):
                self.move_direction = 0
                self.move_das = 0.0
                self.move_arr = 0.0
            else:
                if self.move_das > 0:
                    self.move_das = max(0.0, self.move_das - dt)
                else:
                    self.move_arr -= dt
                    if self.move_arr <= 0:
                        if self.move(self.move_direction, 0):
                            self.move_arr = self.move_repeat_delay
                        else:
                            self.move_arr = self.move_repeat_delay * 1.5

        # Визуальное сглаживание: логика остаётся по клеткам, но отрисовка плавно догоняет.
        self.visual_x = lerp(self.visual_x, float(self.current.x), min(1.0, dt * 18))
        self.visual_y = lerp(self.visual_y, float(self.current.y), min(1.0, dt * 22))

        delay = 0.035 if self.soft_drop else self.fall_delay()
        self.drop_timer += dt

        while self.drop_timer >= delay:
            self.drop_timer -= delay
            if not self.move(0, 1):
                self.lock_timer += delay
                if self.lock_timer > 0.22:
                    self.lock_timer = 0
                    self.lock_piece()
                break
            else:
                self.lock_timer = 0
                if self.soft_drop:
                    self.score += 1

    def draw_block(self, surf, x, y, color, size=TETRIS_CELL, alpha=255, ghost=False):
        block = pygame.Surface((size, size), pygame.SRCALPHA)
        c = (*color, alpha)
        dark = tuple(max(0, int(v * 0.45)) for v in color)
        light = tuple(min(255, int(v * 1.25)) for v in color)

        pygame.draw.rect(block, c, (1, 1, size - 2, size - 2), border_radius=6)
        pygame.draw.rect(block, (*light, alpha), (4, 4, size - 8, max(3, size // 5)), border_radius=4)
        pygame.draw.rect(block, (*dark, alpha), (1, 1, size - 2, size - 2), 2, border_radius=6)

        if ghost:
            block.set_alpha(75)

        surf.blit(block, (x, y))

    def draw_piece_preview(self, surf, piece, ox, oy, title):
        draw_text(surf, self.font_small, title, ox, oy, MUTED)
        if piece is None:
            draw_text(surf, self.font, "—", ox + 22, oy + 38, MUTED)
            return

        mat = piece.matrix
        size = 18
        color = piece.color
        px = ox + 20
        py = oy + 28

        for r, row in enumerate(mat):
            for c, val in enumerate(row):
                if val:
                    self.draw_block(surf, px + c * size, py + r * size, color, size=size, alpha=220)

    def draw(self, surf):
        # Затемнённый космический фон поверх общего меню.
        overlay = pygame.Surface((WIDTH, HEIGHT), pygame.SRCALPHA)
        overlay.fill((3, 8, 19, 120))
        surf.blit(overlay, (0, 0))

        board_rect = pygame.Rect(TETRIS_X - 10, TETRIS_Y - 10, TETRIS_COLS * TETRIS_CELL + 20, TETRIS_ROWS * TETRIS_CELL + 20)
        pygame.draw.rect(surf, (10, 18, 38), board_rect, border_radius=18)
        pygame.draw.rect(surf, (55, 120, 220), board_rect, 2, border_radius=18)

        # Сетка
        for r in range(TETRIS_ROWS):
            for c in range(TETRIS_COLS):
                x = TETRIS_X + c * TETRIS_CELL
                y = TETRIS_Y + r * TETRIS_CELL
                pygame.draw.rect(surf, (16, 27, 55), (x, y, TETRIS_CELL - 1, TETRIS_CELL - 1), border_radius=4)

        # Заполненные клетки
        for r in range(TETRIS_ROWS):
            for c in range(TETRIS_COLS):
                kind = self.board[r][c]
                if kind:
                    self.draw_block(surf, TETRIS_X + c * TETRIS_CELL, TETRIS_Y + r * TETRIS_CELL, TETRIS_COLORS[kind])

        # Подсветка очищаемых линий
        for row, left in self.line_flash:
            a = int(210 * clamp(left / 0.28, 0, 1))
            flash = pygame.Surface((TETRIS_COLS * TETRIS_CELL, TETRIS_CELL), pygame.SRCALPHA)
            flash.fill((*WHITE, a))
            surf.blit(flash, (TETRIS_X, TETRIS_Y + row * TETRIS_CELL))

        # Призрак
        if not self.game_over:
            gy = self.ghost_y()
            for r, row in enumerate(self.current.matrix):
                for c, val in enumerate(row):
                    if val and gy + r >= 0:
                        self.draw_block(
                            surf,
                            TETRIS_X + (self.current.x + c) * TETRIS_CELL,
                            TETRIS_Y + (gy + r) * TETRIS_CELL,
                            self.current.color,
                            alpha=80,
                            ghost=True
                        )

            # Текущая фигура
            bob = math.sin(self.t * 10) * 1.4
            for r, row in enumerate(self.current.matrix):
                for c, val in enumerate(row):
                    if val and self.current.y + r >= 0:
                        self.draw_block(
                            surf,
                            TETRIS_X + (self.visual_x + c) * TETRIS_CELL,
                            TETRIS_Y + (self.visual_y + r) * TETRIS_CELL + bob,
                            self.current.color
                        )

        # Частицы
        for p in self.particles:
            p.draw(surf)

        # Панели
        left_panel = pygame.Rect(36, 70, 180, 355)
        right_panel = pygame.Rect(560, 70, 350, 355)
        for panel in [left_panel, right_panel]:
            pygame.draw.rect(surf, (10, 18, 38), panel, border_radius=18)
            pygame.draw.rect(surf, (55, 120, 220), panel, 1, border_radius=18)

        draw_text(surf, self.font_big, "NEON TETRIS", WIDTH // 2, 22, CYAN, center=True)

        self.draw_piece_preview(surf, self.hold, 58, 96, "HOLD / C")
        draw_text(surf, self.font_small, "Счёт", 58, 210, MUTED)
        draw_text(surf, self.font_mid, str(self.score), 58, 232, WHITE)
        draw_text(surf, self.font_small, "Линии", 58, 282, MUTED)
        draw_text(surf, self.font_mid, str(self.lines), 58, 304, WHITE)
        draw_text(surf, self.font_small, "Уровень", 58, 354, MUTED)
        draw_text(surf, self.font_mid, str(self.level), 58, 376, YELLOW)

        y = 96
        for i, piece in enumerate(self.next_items):
            self.draw_piece_preview(surf, piece, 590, y, "NEXT" if i == 0 else "")
            y += 90

        draw_text(surf, self.font_small, "←/→ или A/D — двигать", 590, 355, MUTED)
        draw_text(surf, self.font_small, "↑/W/X — поворот", 590, 377, MUTED)
        draw_text(surf, self.font_small, "Space — сбросить", 590, 399, MUTED)
        draw_text(surf, self.font_small, "C/Shift — hold   Esc — выбор игры", 590, 421, MUTED)

        if self.message_timer > 0:
            k = clamp(self.message_timer, 0, 1)
            draw_text(surf, self.font_big, self.message, WIDTH // 2, 500 - (1-k) * 20, YELLOW, center=True)

        if self.pause:
            self.draw_center_panel(surf, "ПАУЗА", "P — продолжить   Esc — выбор игры")

        if self.game_over:
            self.draw_center_panel(surf, "TETRIS OVER", "Enter / Space — заново   Esc — выбор игры")

    def draw_center_panel(self, surf, title, subtitle):
        panel = pygame.Rect(250, 200, 460, 145)
        glass = pygame.Surface((panel.w, panel.h), pygame.SRCALPHA)
        glass.fill((8, 16, 36, 230))
        surf.blit(glass, panel)
        pygame.draw.rect(surf, (60, 120, 255), panel, 2, border_radius=22)
        draw_text(surf, self.font_big, title, panel.centerx, panel.y + 42, CYAN, center=True)
        draw_text(surf, self.font, subtitle, panel.centerx, panel.y + 98, MUTED, center=True)


# ============================================================
# ОСНОВНОЙ КЛАСС GAME — ИМЯ НЕ МЕНЯЕМ
# ============================================================

class Game:
    def __init__(self):
        pygame.init()
        pygame.display.set_caption("IFC CHECKER — Game Center")
        self.screen = pygame.display.set_mode((WIDTH, HEIGHT), pygame.DOUBLEBUF)
        self.clock = pygame.time.Clock()

        self.font_big = pygame.font.SysFont("Segoe UI", 44, bold=True)
        self.font_mid = pygame.font.SysFont("Segoe UI", 24, bold=True)
        self.font = pygame.font.SysFont("Segoe UI", 18)
        self.font_small = pygame.font.SysFont("Segoe UI", 14)

        self.surface = pygame.Surface((WIDTH, HEIGHT), pygame.SRCALPHA)
        self.username = get_windows_user()
        self.leaderboard = load_leaderboard()
        self.score_saved = False

        self.mode = None
        self.select_index = 0
        self.select_cards = []
        self.tetris = TetrisGame(self)

        self.reset()

    def reset(self):
        self.player = Player()
        self.enemies = []
        self.particles = []
        self.orbs = []
        self.score = 0
        self.best = max([s for _, s in self.leaderboard], default=0)
        self.t = 0
        self.wave = 1
        self.spawn_timer = 0.5
        self.shot_timer = 0
        self.state = "select"
        self.shake = 0
        self.combo = 1
        self.combo_timer = 0
        self.score_saved = False
        self.gameover_timer = 0
        self.shake_time = 0
        self.shake_duration = 1.5

        self.stars = []
        for _ in range(140):
            self.stars.append([
                random.uniform(0, WIDTH),
                random.uniform(0, HEIGHT),
                random.uniform(0.2, 1.0),
                random.choice([CYAN, BLUE, PURPLE, WHITE])
            ])

    def start_neon(self):
        self.reset_neon_round()
        self.mode = "neon"
        self.state = "neon_menu"

    def reset_neon_round(self):
        self.player = Player()
        self.enemies = []
        self.particles = []
        self.orbs = []
        self.score = 0
        self.best = max([s for _, s in self.leaderboard], default=0)
        self.wave = 1
        self.spawn_timer = 0.5
        self.shot_timer = 0
        self.shake = 0
        self.combo = 1
        self.combo_timer = 0
        self.score_saved = False
        self.gameover_timer = 0
        self.shake_time = 0
        self.shake_duration = 1.5
        if hasattr(self, "laser"):
            del self.laser

    def start_tetris(self):
        self.mode = "tetris"
        self.tetris.reset()
        self.state = "tetris"

    def spawn_enemy(self):
        side = random.randint(0, 3)
        margin = 50
        if side == 0:
            x, y = random.uniform(0, WIDTH), -margin
        elif side == 1:
            x, y = WIDTH + margin, random.uniform(0, HEIGHT)
        elif side == 2:
            x, y = random.uniform(0, WIDTH), HEIGHT + margin
        else:
            x, y = -margin, random.uniform(0, HEIGHT)

        kind = random.choices(
            ["chaser", "speeder", "tank"],
            weights=[58, 30 + self.wave, max(5, self.wave * 2)],
            k=1
        )[0]
        self.enemies.append(Enemy(kind, x, y))

    def shoot(self):
        mx, my = pygame.mouse.get_pos()
        dx, dy = normalize(mx - self.player.x, my - self.player.y)

        if dx == 0 and dy == 0:
            return

        self.shot_timer = 0.095

        hit_enemy = None
        best_proj = 999999

        for enemy in self.enemies:
            ex = enemy.x - self.player.x
            ey = enemy.y - self.player.y
            proj = ex * dx + ey * dy

            if proj < 0:
                continue

            closest_x = self.player.x + dx * proj
            closest_y = self.player.y + dy * proj
            dist = length(enemy.x - closest_x, enemy.y - closest_y)

            if dist < enemy.radius + 8 and proj < best_proj:
                best_proj = proj
                hit_enemy = enemy

        end_x = self.player.x + dx * 760
        end_y = self.player.y + dy * 760

        if hit_enemy:
            end_x, end_y = hit_enemy.x, hit_enemy.y
            hit_enemy.hit(14)

            for _ in range(12):
                a = random.uniform(0, math.tau)
                s = random.uniform(80, 330)
                self.particles.append(Particle(
                    hit_enemy.x, hit_enemy.y,
                    math.cos(a) * s,
                    math.sin(a) * s,
                    random.uniform(2, 6),
                    hit_enemy.color,
                    random.uniform(0.15, 0.38),
                    0.38
                ))

            if hit_enemy.dead:
                self.score += int(hit_enemy.value * self.combo)
                self.combo = min(8, self.combo + 0.25)
                self.combo_timer = 2.1
                self.shake = max(self.shake, 3.5)

                if random.random() < 0.75:
                    self.orbs.append(Orb(hit_enemy.x, hit_enemy.y))

                for _ in range(24):
                    a = random.uniform(0, math.tau)
                    s = random.uniform(120, 480)
                    self.particles.append(Particle(
                        hit_enemy.x, hit_enemy.y,
                        math.cos(a) * s,
                        math.sin(a) * s,
                        random.uniform(2, 8),
                        random.choice([hit_enemy.color, CYAN, WHITE]),
                        random.uniform(0.22, 0.6),
                        0.6
                    ))

        self.laser = (self.player.x, self.player.y, end_x, end_y, 0.045)

    def update(self, dt):
        self.t += dt

        if self.state == "tetris":
            self.tetris.update(dt)
            return

        if self.state == "gameover":
            self.shake_time += dt

        if self.state != "play":
            if self.state == "gameover":
                self.gameover_timer += dt

            for p in list(self.particles):
                if not p.update(dt):
                    self.particles.remove(p)
            return

        keys = pygame.key.get_pressed()
        mouse = pygame.mouse.get_pressed()

        self.player.update(dt, keys, mouse)

        if mouse[0] and self.shot_timer <= 0:
            self.shoot()

        if keys[pygame.K_LSHIFT] or keys[pygame.K_RSHIFT] or mouse[2]:
            self.player.dash(self.particles)

        self.shot_timer = max(0, self.shot_timer - dt)

        self.spawn_timer -= dt
        self.wave = 1 + self.score // 250

        if self.spawn_timer <= 0:
            self.spawn_enemy()
            self.spawn_timer = max(0.18, 0.95 - self.wave * 0.045)

        while len(self.enemies) < min(12 + self.wave * 2, 7 + self.wave):
            if random.random() < 0.015:
                self.spawn_enemy()
            else:
                break

        for enemy in list(self.enemies):
            enemy.update(dt, self.player, self.enemies)

            if length(enemy.x - self.player.x, enemy.y - self.player.y) < enemy.radius + self.player.radius:
                if self.player.damage(13 if enemy.kind != "tank" else 22, self.particles):
                    self.shake = max(self.shake, 8)
                    enemy.hit(999)

            if enemy.dead:
                self.enemies.remove(enemy)

        for orb in list(self.orbs):
            if orb.update(dt, self.player):
                self.score += 8
                self.player.energy = min(self.player.max_energy, self.player.energy + 18)
                self.player.hp = min(self.player.max_hp, self.player.hp + 3)
                self.orbs.remove(orb)

        for p in list(self.particles):
            if not p.update(dt):
                self.particles.remove(p)

        self.combo_timer -= dt
        if self.combo_timer <= 0:
            self.combo = 1

        self.shake *= 0.88

        if self.player.hp <= 0:
            self.game_over()

    def game_over(self):
        if self.score_saved:
            return

        self.score_saved = True

        try:
            save_score_to_excel(self.username, self.score)
            self.leaderboard = load_leaderboard()
            self.best = max([s for _, s in self.leaderboard], default=self.score)
        except Exception:
            pass

        self.state = "gameover"
        self.shake = 18
        self.shake_time = 0

        for _ in range(120):
            a = random.uniform(0, math.tau)
            s = random.uniform(100, 700)
            self.particles.append(Particle(
                self.player.x, self.player.y,
                math.cos(a) * s,
                math.sin(a) * s,
                random.uniform(2, 9),
                random.choice([RED, PINK, CYAN, WHITE]),
                random.uniform(0.35, 1.0),
                1.0
            ))

    def draw_background(self, surf):
        surf.fill(BG)

        for star in self.stars:
            x, y, z, color = star
            sx = (x + math.sin(self.t * z) * 12) % WIDTH
            sy = (y + self.t * 12 * z) % HEIGHT
            c = tuple(int(v * z) for v in color)
            pygame.draw.circle(surf, c, (int(sx), int(sy)), max(1, int(2 * z)))

        grid_color = (18, 34, 70)
        offset = (self.t * 30) % 40

        for x in range(-40, WIDTH + 40, 40):
            pygame.draw.line(surf, grid_color, (x, 0), (x + 120, HEIGHT), 1)

        for y in range(int(offset), HEIGHT, 40):
            pygame.draw.line(surf, grid_color, (0, y), (WIDTH, y), 1)

        pygame.draw.circle(surf, (9, 25, 52), (WIDTH // 2, HEIGHT // 2), 230, 2)
        pygame.draw.circle(surf, (12, 30, 65), (WIDTH // 2, HEIGHT // 2), 150, 1)

    def draw_ui(self, surf):
        if self.state in ("select", "neon_menu", "tetris"):
            return

        draw_text(surf, self.font_mid, f"SCORE {self.score}", 22, 18, WHITE)
        draw_text(surf, self.font_small, f"BEST {self.best}", 24, 48, MUTED)
        draw_text(surf, self.font_small, f"USER {self.username}", 24, 68, MUTED)
        draw_text(surf, self.font_small, f"WAVE {self.wave}", WIDTH - 95, 22, MUTED)

        self.bar(surf, 22, HEIGHT - 38, 210, 12, self.player.hp / self.player.max_hp, RED, "HP")
        self.bar(surf, 252, HEIGHT - 38, 210, 12, self.player.energy / self.player.max_energy, CYAN, "ENERGY")

        if self.combo > 1:
            draw_text(surf, self.font_mid, f"x{self.combo:.1f}", WIDTH // 2, 26, YELLOW, center=True)

        draw_text(
            surf,
            self.font_small,
            "WASD/стрелки — движение   ЛКМ — стрелять   Shift/ПКМ — рывок   Esc — выбор игры",
            WIDTH // 2,
            HEIGHT - 18,
            MUTED,
            center=True
        )

    def bar(self, surf, x, y, w, h, k, color, label):
        k = clamp(k, 0, 1)
        pygame.draw.rect(surf, (22, 32, 55), (x, y, w, h), border_radius=8)
        pygame.draw.rect(surf, color, (x, y, int(w * k), h), border_radius=8)
        pygame.draw.rect(surf, (65, 85, 125), (x, y, w, h), 1, border_radius=8)
        draw_text(surf, self.font_small, label, x, y - 18, MUTED)

    def draw_select(self, surf):
        draw_text(surf, self.font_big, "IFC GAME CENTER", WIDTH // 2, 72, CYAN, center=True)
        draw_text(surf, self.font, "Выбери игру на время проверки моделей", WIDTH // 2, 120, MUTED, center=True)

        cards = [
            {
                "title": "NEON CORE",
                "subtitle": "Аркадный шутер с таблицей лидеров",
                "accent": CYAN,
                "rect": pygame.Rect(105, 175, 340, 250),
                "mode": "neon",
                "keys": "WASD • ЛКМ • Shift",
            },
            {
                "title": "NEON TETRIS",
                "subtitle": "Красивый тетрис без сохранения статистики",
                "accent": YELLOW,
                "rect": pygame.Rect(515, 175, 340, 250),
                "mode": "tetris",
                "keys": "← → ↑ ↓ • Space • C",
            },
        ]
        self.select_cards = cards

        mx, my = pygame.mouse.get_pos()

        for i, card in enumerate(cards):
            rect = card["rect"]
            hovered = rect.collidepoint(mx, my)
            selected = i == self.select_index
            accent = card["accent"]

            pulse = (math.sin(self.t * 4 + i) + 1) / 2
            scale_glow = 1 if selected or hovered else 0.35

            glow = pygame.Surface((rect.w + 60, rect.h + 60), pygame.SRCALPHA)
            pygame.draw.rect(
                glow,
                (*accent, int(45 * scale_glow + pulse * 25 * scale_glow)),
                glow.get_rect(),
                border_radius=34
            )
            surf.blit(glow, (rect.x - 30, rect.y - 30), special_flags=pygame.BLEND_PREMULTIPLIED)

            pygame.draw.rect(surf, (10, 18, 38), rect, border_radius=26)
            pygame.draw.rect(surf, accent if selected or hovered else (50, 75, 115), rect, 3 if selected else 1, border_radius=26)

            # Мини-иконка
            icon_center = (rect.centerx, rect.y + 72)
            if card["mode"] == "neon":
                pygame.draw.circle(surf, (20, 80, 150), icon_center, 35)
                pygame.draw.circle(surf, CYAN, icon_center, 22)
                for a in [0, 2.1, 4.2]:
                    ex = icon_center[0] + math.cos(a + self.t) * 70
                    ey = icon_center[1] + math.sin(a + self.t) * 34
                    pygame.draw.circle(surf, RED, (int(ex), int(ey)), 10)
            else:
                bx = rect.centerx - 42
                by = rect.y + 42
                preview = [("T", 1, 0), ("I", 0, 1), ("O", 2, 1), ("S", 1, 2)]
                for kind, cx, cy in preview:
                    color = TETRIS_COLORS[kind]
                    pygame.draw.rect(surf, color, (bx + cx * 28, by + cy * 28, 25, 25), border_radius=6)
                    pygame.draw.rect(surf, WHITE, (bx + cx * 28 + 4, by + cy * 28 + 4, 17, 5), border_radius=3)

            draw_text(surf, self.font_mid, card["title"], rect.centerx, rect.y + 128, WHITE, center=True)
            draw_text(surf, self.font_small, card["subtitle"], rect.centerx, rect.y + 164, MUTED, center=True)
            draw_text(surf, self.font_small, card["keys"], rect.centerx, rect.y + 198, accent, center=True)

            if selected:
                draw_text(surf, self.font_small, "ENTER — запустить", rect.centerx, rect.y + 225, WHITE, center=True)

        draw_text(surf, self.font_small, "← / → — выбрать     Enter / клик — запустить     Esc — закрыть окно игры", WIDTH // 2, 505, MUTED, center=True)

    def draw_neon_menu(self, surf):
        draw_text(surf, self.font_big, "NEON CORE", WIDTH // 2, 155, CYAN, center=True)
        draw_text(surf, self.font, "Шутер на выживание с таблицей лидеров", WIDTH // 2, 205, MUTED, center=True)
        draw_text(surf, self.font_mid, "Пробел / Enter / ЛКМ — начать", WIDTH // 2, 270, WHITE, center=True)
        draw_text(surf, self.font_small, f"Текущий пользователь: {self.username}", WIDTH // 2, 310, MUTED, center=True)
        draw_text(surf, self.font_small, "Esc — назад к выбору игры", WIDTH // 2, 342, MUTED, center=True)

    def draw_leaderboard(self, surf):
        panel = pygame.Rect(270, 105, 420, 350)
        pygame.draw.rect(surf, (12, 20, 40), panel, border_radius=18)
        pygame.draw.rect(surf, (60, 120, 255), panel, 2, border_radius=18)

        draw_text(surf, self.font_mid, "ТАБЛИЦА ЛИДЕРОВ", WIDTH // 2, 130, CYAN, center=True)

        rows = self.leaderboard[:10]
        if not rows:
            draw_text(surf, self.font, "Пока нет результатов", WIDTH // 2, 260, MUTED, center=True)
            return

        y = 175
        for i, (user, score) in enumerate(rows, start=1):
            color = YELLOW if user.lower() == self.username.lower() else WHITE
            draw_text(surf, self.font, f"{i}. {user}", 305, y, color)
            draw_text(surf, self.font, str(score), 650, y, color, center=False)
            y += 25

    def draw_gameover(self, surf):
        draw_text(surf, self.font_big, "СИСТЕМА ПОВРЕЖДЕНА", WIDTH // 2, 58, RED, center=True)
        draw_text(surf, self.font_mid, f"Твои очки: {self.score}", WIDTH // 2, 92, WHITE, center=True)
        self.draw_leaderboard(surf)
        if self.gameover_timer < 3.0:
            wait_text = f"Новая попытка будет доступна через {3.0 - self.gameover_timer:.1f} сек."
            draw_text(surf, self.font, wait_text, WIDTH // 2, 490, MUTED, center=True)
        else:
            draw_text(surf, self.font, "Enter / ЛКМ — новая попытка     Esc — выбор игры", WIDTH // 2, 490, MUTED, center=True)

    def draw(self):
        ox = oy = 0

        if self.state == "gameover":
            if self.shake_time < self.shake_duration:
                k = 1 - self.shake_time / self.shake_duration
                strength = max(1, int(14 * k))
                ox = random.randint(-strength, strength)
                oy = random.randint(-strength, strength)
        else:
            if self.shake > 0.4:
                ox = random.randint(-int(self.shake), int(self.shake))
                oy = random.randint(-int(self.shake), int(self.shake))

        self.surface.fill((0, 0, 0, 0))
        self.draw_background(self.surface)

        if self.state == "tetris":
            self.tetris.draw(self.surface)
        else:
            for orb in self.orbs:
                orb.draw(self.surface, self.t)

            for enemy in self.enemies:
                enemy.draw(self.surface, self.t)

            for p in self.particles:
                p.draw(self.surface)

            if self.state == "play":
                self.player.draw(self.surface, self.t)

            if hasattr(self, "laser"):
                x1, y1, x2, y2, life = self.laser
                life -= 1 / FPS
                if life > 0:
                    self.laser = (x1, y1, x2, y2, life)
                    pygame.draw.line(self.surface, (150, 240, 255), (x1, y1), (x2, y2), 5)
                    pygame.draw.line(self.surface, WHITE, (x1, y1), (x2, y2), 2)
                else:
                    del self.laser

            self.draw_ui(self.surface)

            if self.state == "select":
                self.draw_select(self.surface)
            elif self.state == "neon_menu":
                self.draw_neon_menu(self.surface)
            elif self.state == "gameover":
                self.draw_gameover(self.surface)

        self.screen.fill(BG)
        self.screen.blit(self.surface, (ox, oy))
        pygame.display.flip()

    def handle_events(self):
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                sys.exit()

            if self.state == "tetris":
                self.tetris.handle_event(event)
                continue

            if event.type == pygame.KEYDOWN:
                if self.state == "select":
                    if event.key in (pygame.K_LEFT, pygame.K_a):
                        self.select_index = (self.select_index - 1) % 2
                    elif event.key in (pygame.K_RIGHT, pygame.K_d):
                        self.select_index = (self.select_index + 1) % 2
                    elif event.key in (pygame.K_RETURN, pygame.K_SPACE):
                        if self.select_index == 0:
                            self.start_neon()
                        else:
                            self.start_tetris()
                    elif event.key == pygame.K_ESCAPE:
                        pygame.quit()
                        sys.exit()
                    continue

                if event.key == pygame.K_ESCAPE:
                    if self.state in ("play", "neon_menu", "gameover"):
                        self.state = "select"
                        continue

                if event.key in (pygame.K_SPACE, pygame.K_RETURN):
                    if self.state == "neon_menu":
                        self.state = "play"
                    elif self.state == "gameover":
                        if self.gameover_timer >= 3.0:
                            self.reset_neon_round()
                            self.state = "play"

            if event.type == pygame.MOUSEBUTTONDOWN:
                if self.state == "select":
                    mx, my = event.pos
                    for i, card in enumerate(self.select_cards):
                        if card["rect"].collidepoint(mx, my):
                            self.select_index = i
                            if card["mode"] == "neon":
                                self.start_neon()
                            else:
                                self.start_tetris()
                            break
                    continue

                if self.state == "neon_menu":
                    self.state = "play"
                elif self.state == "gameover":
                    if self.gameover_timer >= 3.0:
                        self.reset_neon_round()
                        self.state = "play"

    def run(self):
        while True:
            dt = self.clock.tick(FPS) / 1000
            dt = min(dt, 1 / 45)

            self.handle_events()
            self.update(dt)
            self.draw()


if __name__ == "__main__":
    Game().run()
